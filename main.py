from pathlib import Path
from openpyxl import Workbook
from datetime import date, datetime
from openpyxl.styles import Font, PatternFill, Side, Border, Alignment
import openpyxl
import os
import math
import shutil

today = date.today()
current_year = today.year
excel_name = f'{current_year}-'

green_fill = PatternFill(start_color='009E4D', end_color='009E4D', fill_type='solid')
center_alignment = Alignment(horizontal='center')
bold_font = Font(bold=True, size=12)

raw_directory = Path.cwd() /"resources"/"raw"
processed_directory = Path.cwd() /"resources"/"processed"
current_directory = Path.cwd()

excel_files_xlsx = list(raw_directory.glob('*.xlsx'))
excel_files_xls = list(raw_directory.glob('*.xls'))
all_excel_files = excel_files_xlsx + excel_files_xls

def exec():
    print(f'[LOG] Total files: {all_excel_files}')

    process(all_excel_files)

def process(all_excel_files):
    print(f'[LOG] Starting to process the files: {all_excel_files}')
    rawData = [[0 for x in range(4)] for y in range(len(all_excel_files))]

    for i in range(len(all_excel_files)):
        file_path = all_excel_files[i]

        print(f"[LOG] Processing file: {file_path}")
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook['Technical']

        name = sheet.cell(row=3, column=3).value
        id = sheet.cell(row=4, column=3).value
        score = sheet.cell(row=3, column=5).value
        date = sheet.cell(row=6, column=3).value

        rawData[i] = [name, id, score, date]
        workbook.close()

    processToReport(rawData, all_excel_files)

def processToReport(rawData, all_excel_files):
    for row in range(len(rawData)):
        week_of_month = get_week_of_month(rawData[row][3])
        month = rawData[row][3].month
        createFile(month)

        file_path_month = current_directory / f'{excel_name}{month}.xlsx'
        wb = openpyxl.load_workbook(file_path_month)
        ws = wb['Quality Audit']

        rowNum = find_next_last_row_with_data(ws, 'A')

        createWorksheet(ws, rowNum, 1, rawData[row][1])
        createWorksheet(ws, rowNum, 2, rawData[row][0])
        createWorksheet(ws, rowNum, week_of_month+2, rawData[row][2])

        wb.save(file_path_month)
        wb.close()

        move(all_excel_files[row])

def move(source_file):
    os.makedirs(processed_directory, exist_ok=True)
    shutil.move(source_file, processed_directory)

def find_next_last_row_with_data(worksheet, column_letter):
    for row in range(worksheet.max_row, 0, -1):
        if worksheet[f'{column_letter}{row}'].value is not None:
            return row+1
    return 0

def createWb(monthNum):
    wb = Workbook()
    ws = wb.active

    ws.title = "Dashboard"
    createWorksheetHeader(ws, 1, "Employee ID")
    createWorksheetHeader(ws, 2, "Employee Name")
    createWorksheetHeader(ws, 3, "Quality Audit")
    createWorksheetHeader(ws, 4, "Customer Satisfaction")
    createWorksheetHeader(ws, 5, "MLL Hours")
    createWorksheetHeader(ws, 6, "Total Scores")

    ws2 = wb.create_sheet('Quality Audit')
    createWorksheetHeader(ws2, 1, "Employee ID")
    createWorksheetHeader(ws2, 2, "Name")
    createWorksheetHeader(ws2, 3, "W1")
    createWorksheetHeader(ws2, 4, "W2")
    createWorksheetHeader(ws2, 5, "W3")
    createWorksheetHeader(ws2, 6, "W4")
    createWorksheetHeader(ws2, 7, "W5")
    createWorksheetHeader(ws2, 8, "W6")
    createWorksheetHeader(ws2, 9, "W7")

    ws3 = wb.create_sheet('Customer Satisfaction')
    createWorksheetHeader(ws3, 1, "Employee ID")
    createWorksheetHeader(ws3, 2, "Name")
    createWorksheetHeader(ws3, 3, "W1")
    createWorksheetHeader(ws3, 4, "W2")
    createWorksheetHeader(ws3, 5, "W3")
    createWorksheetHeader(ws3, 6, "W4")
    createWorksheetHeader(ws3, 7, "W5")
    createWorksheetHeader(ws3, 8, "W6")
    createWorksheetHeader(ws3, 9, "W7")


    wb.save(f'{excel_name}{monthNum}.xlsx')
    wb.close()
    print(f"[LOG] Successfully created the excel file: {excel_name}.xlsx")

def createWorksheetHeader(ws, col, val):
    ws.cell(row=1, column=col, value=val).fill = green_fill

def createWorksheet(ws, row, col, val):
    ws.cell(row=row, column=col, value=val)

def get_week_of_month(date_obj):
    first_day = date_obj.replace(day=1)
    dom = date_obj.day

    adjusted_dom = dom + first_day.weekday()
    return int(math.ceil(adjusted_dom / 7.0))

def createFile(monthNum):
    file_path_month = current_directory / f'{excel_name}{monthNum}.xlsx'
    if os.path.isfile(file_path_month):
        print(f"[LOG] The file '{file_path_month}' exists.")

    else:
        print(f"[LOG] The file '{file_path_month}' does not exist.")
        createWb(monthNum)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    exec()
